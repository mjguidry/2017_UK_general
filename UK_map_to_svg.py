# -*- coding: utf-8 -*-
"""
Created on Sat Apr 01 14:55:06 2017

@author: MGuidry
"""

import tempfile

temp_dir=tempfile.tempdir

# This is the input sheet
tmp_file=r'UK_block_map.xlsx'
sheet='GB'
start_cell='I6'
end_cell='AJ56'
addText=False
joinCells=True

# This is the input file
im_file=temp_dir+'/tmp.png'

# This is the output PNG file, which will be in black and white
save_file='UK_BW.png'

# This is the SVG file
svg_file=r'UK.svg'

# This is the CSV file
csv_file=r'UK_coords.csv'

# Imports
from openpyxl import load_workbook
from PIL import Image,ImageColor
from numpy import zeros
import cv2
import base64
import os, re
import csv

#Dimensions, in pixels
xw=16
yw=16

wb=load_workbook(filename=tmp_file)
ws=wb[sheet]
rng=tuple(ws[start_cell:end_cell])
nrows=len(rng)
ncols=len(rng[0])
comment_dict=dict()
if addText:
    text_dict=dict()
    text_cnt=0
im=Image.new("RGB",(ncols*(xw+1)+1,nrows*(yw+1)+1),"white")
pix=im.load()
for col in range(ncols):
    for row in range(nrows):
        color=rng[row][col].fill.bgColor.rgb
        if(color[0:8]=='00000000'):
            R=255
            G=255
            B=255
        else:
            R=int(color[2:4],16)
            G=int(color[4:6],16)
            B=int(color[6:8],16)
        for x in range(xw):
            for y in range(yw):
                pix[x+1+(xw+1)*col,y+1+(yw+1)*row]=(R,G,B)
        if(rng[row][col].comment):
            key=rng[row][col].comment.text.split('\n')[0]
            key=re.sub(' ','_',key)
            comment_dict[key]=((xw/2+1)+(xw+1)*col,(yw/2+1)+(yw+1)*row)
        if(addText):
            if(rng[row][col].value):
                text_dict[text_cnt]=dict()
                try:
                    if(rng[row][col].value==int(rng[row][col].value)):
                        text_dict[text_cnt]['value']=int(rng[row][col].value)
                    else:
                        text_dict[text_cnt]['value']=rng[row][col].value
                except:
                    text_dict[text_cnt]['value']=rng[row][col].value
                text_dict[text_cnt]['position']=((xw/2+1)+(xw+1)*col,(yw/2+1)+(yw+1)*row)
                text_cnt=text_cnt+1
im.show()
mode=im.mode
size=im.size

xsize=im.size[0]
ysize=im.size[1]
white=ImageColor.getcolor('white',mode)
black=ImageColor.getcolor('black',mode)

color=white

if(joinCells):
    for k in range(2):
        for x in range(xsize):
            for y in range(ysize):
                if(pix[x,y]==white):
                    try:
                        top_color=pix[x,y-1]
                    except:
                        top_color=white
                    try:
                        bot_color=pix[x,y+1]
                    except:
                        bot_color=white
                    try:
                        left_color=pix[x-1,y]
                    except:
                        left_color=white
                    try:
                        right_color=pix[x+1,y]
                    except:
                        right_color=white
                    if(top_color==bot_color and top_color!=white):
                        pix[x,y]=top_color
                    if(left_color==right_color and left_color!=white):
                        pix[x,y]=left_color
                        
minx=xsize
miny=ysize
maxx=0
maxy=0
for x in range(xsize):
    for y in range(ysize):
        if(pix[x,y]!=white):
            pix[x,y]=black
            minx=min(minx,x)
            miny=min(miny,y)
            maxx=max(maxx,x)
            maxy=max(maxy,y)
            
bw=10
im2=Image.new(mode,((2*bw+1)+maxx-minx,(2*bw+1)+maxy-miny),white)
#im2=Image.new(mode,(21+maxx-minx,42+maxy-miny),white)
pix2=im2.load()
for x in range(1+maxx-minx):
    for y in range(1+maxy-miny):
        pix2[bw+x,bw+y]=pix[minx+x,miny+y]
for comment in comment_dict.keys():
    x=comment_dict[comment][0]-minx+bw
    y=comment_dict[comment][1]-miny+bw
    comment_dict[comment]=(x,y)
if(addText):
    for text in text_dict.keys():
        x=text_dict[text]['position'][0]-minx+bw
        y=text_dict[text]['position'][1]-miny+bw
        text_dict[text]['position']=(x,y)
#        pix2[29+x,10+y]=pix[minx+x,miny+y]

# im.show()
im2.show()
im2.save(save_file)
print 'Orig image size is'+'\t'+str(im.size)
print 'New image size is'+'\t'+str(im2.size)


img = cv2.imread(save_file)
gray = cv2.imread(save_file,0)

ret,thresh = cv2.threshold(gray,127,255,1)

try:
    image, contours, h = cv2.findContours(thresh,1,2)
except:
    contours, h = cv2.findContours(thresh,1,2)
    
print 'Number of paths is'+'\t'+str(len(contours))
if(addText):
    contour_dict=dict()    
    for k,contour in enumerate(contours):
        minx=min([c[0][0] for c in contour])
        maxx=max([c[0][0] for c in contour])
        miny=min([c[0][1] for c in contour])
        maxy=max([c[0][1] for c in contour])
        dist_ctr=zeros((maxx-minx+1,maxy-miny+1))
        for x in range(minx,maxx+1):
            for y in range(miny,maxy+1):
                dist_ctr[x-minx,y-miny]=cv2.pointPolygonTest(contour,(x,y),True)
        minVal, maxVal, minLoc, maxLoc = cv2.minMaxLoc(dist_ctr)
        contour_dict[k]=dict()
        contour_dict[k]['center']=(maxLoc[1]+minx,maxLoc[0]+miny)
        contour_dict[k]['radius']=maxVal
    max_text_length=max([len(str(value['value'])) for value in text_dict.values()])
    min_text_space=min([2*value['radius'] for value in contour_dict.values()])
    font_size=int(min_text_space/max_text_length)
    
with open(save_file, "rb") as image_file:
    encoded_string = base64.b64encode(image_file.read())

dwg=open(svg_file,'wb')
csvfile=open(csv_file, 'wb')
csv_writer=csv.writer(csvfile)

dwg.write('''<?xml version="1.0" encoding="UTF-8" standalone="no"?>
<!-- Created with Inkscape (http://www.inkscape.org/) -->

<svg
   xmlns:dc="http://purl.org/dc/elements/1.1/"
   xmlns:cc="http://creativecommons.org/ns#"
   xmlns:rdf="http://www.w3.org/1999/02/22-rdf-syntax-ns#"
   xmlns:svg="http://www.w3.org/2000/svg"
   xmlns="http://www.w3.org/2000/svg"
   xmlns:xlink="http://www.w3.org/1999/xlink"
   xmlns:sodipodi="http://sodipodi.sourceforge.net/DTD/sodipodi-0.dtd"
   xmlns:inkscape="http://www.inkscape.org/namespaces/inkscape"
   id="svg3039"
   version="1.1"
   inkscape:version="0.48.4 r9939"
   width="'''+str(im2.size[0])+'''"
   height="'''+str(im2.size[1])+'''"
   sodipodi:docname="'''+os.path.basename(svg_file)+'''">
  <metadata
     id="metadata3045">
    <rdf:RDF>
      <cc:Work
         rdf:about="">
        <dc:format>image/svg+xml</dc:format>
        <dc:type
           rdf:resource="http://purl.org/dc/dcmitype/StillImage" />
        <dc:title></dc:title>
      </cc:Work>
    </rdf:RDF>
  </metadata>
  <defs
     id="defs3043" />
  <sodipodi:namedview
     pagecolor="#ffffff"
     bordercolor="#666666"
     borderopacity="1"
     objecttolerance="10"
     gridtolerance="10"
     guidetolerance="10"
     inkscape:pageopacity="0"
     inkscape:pageshadow="2"
     inkscape:window-width="'''+str(im2.size[0]+100)+'''"
     inkscape:window-height="'''+str(im2.size[1]+100)+'''"
     id="namedview3041"
     showgrid="false"
     inkscape:zoom="1"
     inkscape:cx="'''+str(im2.size[0]/2)+'''"
     inkscape:cy="'''+str(im2.size[1]/2)+'''"
     inkscape:window-x="0"
     inkscape:window-y="0"
     inkscape:window-maximized="0"
     inkscape:current-layer="svg3039" />
  <image
     width="'''+str(im2.size[0])+'''"
     height="'''+str(im2.size[1])+'''"
     xlink:href="data:image/png;base64,''')
dwg.write(encoded_string)
dwg.write('''"
     id="image3047"
     x="0"
     y="0" />
''')
for m,contour in enumerate(contours):
    path=''
    for k,point in enumerate(contour):
        if(k==0):
            path=path+'M '
        else:
            path=path+'L '
        path=path+str(point[0][0]+0.5)+' '
        path=path+str(point[0][1]+0.5)+' '
    path=path+'z'
    path_id='path_'+str(m)
    for comment in comment_dict.keys():
        x=comment_dict[comment][0]
        y=comment_dict[comment][1]
        dist=cv2.pointPolygonTest(contour,(x,y),False)
        if(dist>0):
            path_id=comment
    dwg.write('''  <path
     style="fill:#000000"
     d="'''+path+'''"
     id="'''+path_id+'''"
     inkscape:label="" />
''')
    csv_writer.writerow([re.sub('_',' ',path_id),path])
    if(addText):
        for text in text_dict.keys():
            x=text_dict[text]['position'][0]
            y=text_dict[text]['position'][1]
            dist=cv2.pointPolygonTest(contour,(x,y),False)
            if(dist>0):
#                M = cv2.moments(contour)
#                cX =(M["m10"] / M["m00"])
#                cY =(M["m01"] / M["m00"])
                cX=contour_dict[m]['center'][0]
                cY=contour_dict[m]['center'][1]
                dwg.write('''  <text
     x="'''+str(cX)+'''" y="'''+str(cY+font_size/2)+'''" 
     style="font-family: Arial;
            font-size: '''+str(font_size)+'''px;
            stroke: #ffffff;
            fill: #ffffff;">
            <tspan text-anchor="middle">
            '''+str(text_dict[text]['value'])+'''
            </tspan>
     </text>                
''')
        
dwg.write('''</svg>
''')
dwg.close()
csvfile.close()