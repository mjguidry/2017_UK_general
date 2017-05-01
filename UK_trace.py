# -*- coding: utf-8 -*-
"""
Created on Fri Apr 28 22:50:36 2017

@author: mike
"""

from PIL import Image, ImageColor, ImageDraw
#import requests
#from StringIO import StringIO
from math import sqrt, ceil, log
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.cell import get_column_letter

img=Image.open('./uk.gif')
xsize=img.size[0]
ysize=img.size[1]
mode="RGB"
im=img.convert(mode)
black=ImageColor.getcolor('black',mode)
white=ImageColor.getcolor('white',mode)
ImageDraw.floodfill(im,(xsize/2,ysize/2),black)
ImageDraw.floodfill(im,(xsize/2,ysize/4),black)
ImageDraw.floodfill(im,(xsize/2,3*ysize/4),black)
#ImageDraw.floodfill(im,(75,345),black)
pix=im.load()
colors=im.getcolors(1024)
num_black=[color for color in colors if color[1]==black][0][0]
num_black=float(num_black)/21/21
target_blocks=632-2
num_gray=num_black


sar_length=int(ceil(log(target_blocks)/log(2)))+1
passed = 0
times_thru =0
new_xsize=xsize
new_ysize=ysize


threshold=0
new_size=float(target_blocks)/num_gray
new_xsize=int(round(new_xsize*sqrt(new_size)))
new_ysize=int(round(new_ysize*sqrt(new_size)))
print num_gray, new_xsize,new_ysize
im2=im.resize((new_xsize,new_ysize))
pix=im2.load()
num_rows=int(ceil(new_ysize/21))
num_cols=int(ceil(new_xsize/21))
for k in range(sar_length):
    threshold=threshold+0.5/2**k
    num_gray=0
    for row in range(num_rows):
        for col in range(num_cols):
            start_x=21*col
            end_x=21*col+20
            start_y=21*row
            end_y=21*row+20
            im_temp=im2.crop((start_x,start_y,end_x,end_y))
            try:
                colors_temp=im_temp.getcolors()
                num_black_temp=[color for color in colors_temp if color[1]==black][0][0]
                if(num_black_temp>21*21*threshold):
                    num_gray=num_gray+1
            except:
                pass
    print k, num_gray, threshold
    if(num_gray<target_blocks):
        last_under=k
        last_under_threshold=threshold
        last_under_cnt=num_gray
        threshold=threshold-0.5/2**k

num_gray=0
wb=Workbook()
ws = wb.active
grayFill=PatternFill(start_color='FF808080',
                     end_color='FF808080',
                     fill_type='solid')
whiteFill=PatternFill(start_color='FFFFFFFF',
                     end_color='FFFFFFFF',
                     fill_type='solid')
for row in range(num_rows):
    ws.row_dimensions[row+2].height = 18
    for col in range(num_cols):
        ws.column_dimensions[get_column_letter((col+2))].width = 2.43
        start_x=21*col
        end_x=21*col+20
        start_y=21*row
        end_y=21*row+20
        im_temp=im2.crop((start_x,start_y,end_x,end_y))
        try:
            colors_temp=im_temp.getcolors()
            num_black_temp=[color for color in colors_temp if color[1]==black][0][0]
            if(num_black_temp>21*21*threshold):
                ws.cell(row=row+2,column=col+2).fill=grayFill
                #ss_function("setCellGray",[ws_name, sheet_name,row+2,col+2])
                #gray_arr.append([row+2,col+2])
                num_gray=num_gray+1
            else:
                #ss_function("setCellColor",[ws_name, sheet_name,row+2,col+2,"white"])
                ws.cell(row=row+2,column=col+2).fill=whiteFill
                #white_arr.append([row+2,col+2])
        except:
            #ss_function("setCellColor",[ws_name, sheet_name,row+2,col+2,"white"])
            #white_arr.append([row+2,col+2])
            ws.cell(row=row+2,column=col+2).fill=whiteFill
            
wb.save(filename = 'UK_trace.xlsx')